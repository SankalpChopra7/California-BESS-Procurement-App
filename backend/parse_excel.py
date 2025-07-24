import json
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import math

ROOT       = Path(__file__).resolve().parents[1]
EXCEL_PATH = ROOT / 'California_Arizona_Texas_Procurement_Data.xlsx'
DATA_DIR   = ROOT / 'data'
DATA_DIR.mkdir(exist_ok=True)

STATE_COORDS = {
    "California": {"lat": 36.7783, "lon": -119.4179},
    "Arizona":    {"lat": 34.0489, "lon": -111.0937},
    "Texas":      {"lat": 31.9686, "lon":  -99.9018},
}

def safe_value(val):
    # Converts pandas NaN → None
    return None if isinstance(val, float) and math.isnan(val) else val

def _extract_lat_lon(row):
    """Try to pull numeric lat/lon columns from a pandas row."""
    lat = lon = None
    for col in row.index:
        lc = col.lower()
        if lat is None and 'lat' in lc:
            lat = safe_value(row[col])
        if lon is None and ('lon' in lc or 'lng' in lc):
            lon = safe_value(row[col])
    return lat, lon

def parse_suppliers():
    # Use openpyxl for hyperlink extraction so we keep hyperlinks for popups
    wb = load_workbook(EXCEL_PATH, data_only=True)
    suppliers = []

    sheet_cfg = [
        # (sheet name, state key, county column, contact column, map column)
        ("California County Data ", "California",
         "California County", "Contact", "Location on Google Maps"),
        ("Arizona County Data ", "Arizona",
         "Arizona  County", "Company Website", "Location on Google Maps"),
        ("Texas County Data", "Texas",
         "Texas County", None, "Link to Google Maps"),
    ]

    for sheet_name, state_key, county_col, contact_hdr, map_hdr in sheet_cfg:
        ws = wb[sheet_name]
        header = [str(c.value).strip() if c.value else "" for c in ws[1]]
        contact_col = header.index(contact_hdr) + 1 if contact_hdr in header else None
        maps_col = header.index(map_hdr) + 1 if map_hdr in header else None

        # read into pandas (to get text fields)
        df = pd.read_excel(EXCEL_PATH, sheet_name=sheet_name, dtype=str)
        df.rename(columns=lambda c: str(c).strip(), inplace=True)

        for idx, row in df.iterrows():
            contact_cell = ws.cell(row=idx+2, column=contact_col) if contact_col else None
            map_cell     = ws.cell(row=idx+2, column=maps_col) if maps_col else None

            lat, lon = _extract_lat_lon(row)
            if lat is None or lon is None:
                lat = STATE_COORDS[state_key]["lat"]
                lon = STATE_COORDS[state_key]["lon"]

            suppliers.append({
                "state":        state_key,
                "county":       safe_value(row.get(county_col)),
                "service_type": safe_value(row.get("Service Type")),
                "company":      safe_value(row.get("Company Name")),
                "notes":        safe_value(row.get("Notes")),
                "contact_url":  contact_cell.hyperlink.target if contact_cell and contact_cell.hyperlink else None,
                "map_url":      map_cell.hyperlink.target     if map_cell and map_cell.hyperlink     else None,
                "lat":          lat,
                "lon":          lon,
            })

    return suppliers

def parse_projects():
    # (you can add hyperlink logic here if needed)
    wb    = pd.ExcelFile(EXCEL_PATH)
    sites = wb.parse("SOLV BESS Sites").rename(columns=lambda c: c.strip())
    projects = []

    for _, row in sites.iterrows():
        lat, lon = _extract_lat_lon(row)
        if lat is None or lon is None:
            lat = STATE_COORDS["California"]["lat"]
            lon = STATE_COORDS["California"]["lon"]

        projects.append({
            "project":   safe_value(row.get("SOLV BESS PROJECT")),
            "location":  safe_value(row.get("Location")),
            "mw_ac":     safe_value(row.get("MW AC Capacity")),
            "mw_dc":     safe_value(row.get("MW DC Capacity")),
            "client":    safe_value(row.get("Client")),
            "lat":       lat,
            "lon":       lon,
        })

    return projects

def _load_or_cache(name, parse_fn):
    path = DATA_DIR / f"{name}.json"
    if path.exists():
        return json.loads(path.read_text())
    data = parse_fn()
    path.write_text(json.dumps(data, indent=2))
    return data

def load_data():
    return {
        "suppliers": _load_or_cache("suppliers", parse_suppliers),
        "projects":  _load_or_cache("projects",  parse_projects),
    }

DATA = load_data()

def get_suppliers():
    return DATA["suppliers"]

def get_projects():
    return DATA["projects"]


if __name__ == "__main__":
    # Regenerate cache on demand
    DATA_DIR.mkdir(exist_ok=True)
    with open(DATA_DIR / "suppliers.json", "w") as f:
        json.dump(parse_suppliers(), f, indent=2)
    with open(DATA_DIR / "projects.json", "w") as f:
        json.dump(parse_projects(), f, indent=2)

